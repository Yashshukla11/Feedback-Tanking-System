import openpyxl

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write headers to the worksheet
worksheet.cell(row=1, column=1, value="Name")
worksheet.cell(row=1, column=2, value="Phone Number")
worksheet.cell(row=1, column=3, value="Email")
worksheet.cell(row=1, column=4, value="Feedback")

# Loop to collect feedback from users
while True:
    # Prompt the user to enter their name
    name = input("Please enter your name (alphabet only): ")
    if not name.isalpha():
        print("Invalid input. Please enter your name in alphabets only.")
        continue

    # Prompt the user to enter their phone number
    phone = input("Please enter your phone number (numeric only): ")
    if not phone.isnumeric():
        print("Invalid input. Please enter your phone number in numeric form only.")
        continue

    # Prompt the user to enter their email address
    email = input("Please enter your email address: ")

    # Prompt the user to enter their feedback
    feedback = input("Please enter your feedback: ")

    # Write the feedback to the worksheet
    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1, value=name)
    worksheet.cell(row=row, column=2, value=phone)
    worksheet.cell(row=row, column=3, value=email)
    worksheet.cell(row=row, column=4, value=feedback)

    # Ask the user if they want to continue or quit
    choice = input("Thank you for your feedback. Press 'q' to quit or any other key to continue: ")
    if choice.lower() == "q":
        break

# Save the workbook to a file
workbook.save("feedback.xlsx")
print("Feedback saved to feedback.xlsx.")
